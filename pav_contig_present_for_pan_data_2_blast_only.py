#!/usr/bin/env python3
# -*- coding: utf-8 -*-
from ast import Num
from itertools import count
from subprocess import STDOUT
from sys import stderr
from typing import Counter
from gffutils import attributes
from joblib import Parallel,delayed
from Directory_creater import directory_creater
from Bio.Blast.Applications import NcbiblastformatterCommandline, NcbimakeblastdbCommandline
from Bio.Blast.Applications import NcbiblastnCommandline
from Bio.Blast.Applications import NcbimakeblastdbCommandline
from Bio.Blast import NCBIXML
import os 
import re
from openpyxl import Workbook
from pathlib import Path
import subprocess
import gffutils
from Bio.Application import ApplicationError

#output,remmben save at end

# excel_book=openpyxl.load_workbook('../test_why_00005_so_many_1/test.xlsx')

pan_sh=None
dic_gene={}
gene_count=1

def out_to_excel(count,gene,value):
    count=count+1
    global gene_count

    if dic_gene.__contains__(gene):
        pan_sh.cell(dic_gene[gene],count,value)
    else:
        gene_count=gene_count+1
        pan_sh.cell(gene_count,1,'Pan_{}'.format(gene_count-1))
        pan_sh.cell(gene_count,2,gene)

        dic_gene.setdefault(gene,gene_count)
        pan_sh.cell(dic_gene[gene],count,value)

def excel_species_name(count,name):
    pan_sh.cell(1,count,name)

def run_blast(query_file,species_id_path,species_out_path):
    make_db_cmd=NcbimakeblastdbCommandline(
        cmd='/mnt/d/zhes_learning_space/software_in_ubuntu/ncbi-blast-2.10.1+/bin/makeblastdb',
        dbtype='nucl',
        input_file=species_id_path,
        out=str(species_out_path/"blastdb"/species_id_path.stem)
    )
    blast_cmd=NcbiblastnCommandline(
            cmd='/mnt/d/zhes_learning_space/software_in_ubuntu/ncbi-blast-2.10.1+/bin/blastn',
            query=query_file,
            db=species_out_path/"blastdb"/species_id_path.stem,
            outfmt=11,
            out=species_out_path/"asn"/(species_id_path.stem+".asn")
            # perc_identity=95
        )
    blast_xml_cmd=NcbiblastformatterCommandline(
    archive=species_out_path/"asn"/(species_id_path.stem+".asn"),
    outfmt=5,
    out=species_out_path/"xml"/(species_id_path.stem+".xml"),
    cmd='/mnt/d/zhes_learning_space/software_in_ubuntu/ncbi-blast-2.10.1+/bin/blast_formatter'
    )
    blast_txt_cmd=NcbiblastformatterCommandline(
        archive=species_out_path/"asn"/(species_id_path.stem+".asn"),
        outfmt=7,
        out=species_out_path/"txt"/(species_id_path.stem+".txt"),
        cmd='/mnt/d/zhes_learning_space/software_in_ubuntu/ncbi-blast-2.10.1+/bin/blast_formatter'
    )
    db_file=species_out_path/"blastdb"/(species_id_path.stem+".ndb")
    if (species_out_path/"xml"/(species_id_path.stem+".xml")).exists() is False:
        if db_file.exists() is False:
            make_db_cmd()
        try:
            blast_cmd()
        except ApplicationError:
            print(blast_xml_cmd)
        try:
            blast_txt_cmd()
        except ApplicationError:
            print(blast_xml_cmd)
        try:
            blast_xml_cmd()
        except ApplicationError:
            print(blast_xml_cmd)
    # if os.path.getsize(str(species_out_path/"xml"/(species_id_path.stem+".xml"))) == 0:
    #     blast_cmd()
    #     blast_txt_cmd()
    #     blast_xml_cmd()

def run_gth(query_file,species_id_path,species_out_path):
    run_gth_popen=subprocess.Popen(
        [
            "/mnt/d/zhes_learning_space/software_in_ubuntu/gth-1.7.3-Linux_x86_64-64bit/bin/gth",
            "-genomic",
            species_id_path,
            "-cdna",
            query_file,
            "-gff3out",
            "-o",
            species_out_path/"gth_out"/(species_id_path.stem+".gff"),
            "-intermediate",
            "-force"
        ],
        stderr=subprocess.STDOUT,
        stdout=(species_out_path/"gth_stdout"/(species_id_path.stem+".txt")).open("w")
    )
    run_gth_popen.wait()

def update_gff_db(strain_gff,db_out_dir):
    def gff_update_iterator(db):
        for gene in db.features_of_type('gene'):

            # modify attributes

            # add a new attribute for exon id
            for target_id in gene.attributes["Target"][0].split(","):
                gene.attributes['ID'] = target_id.split(" ")[0]+"_"+gene.chrom
                gene.attributes['Target']=target_id
                yield gene

    try:
        gffutils.create_db(
        str(strain_gff),
        dbfn=str(db_out_dir/(strain_gff.stem+".db"))
        )
    except:
        pass
    db = gffutils.FeatureDB(str(db_out_dir/(strain_gff.stem+".db")))
    db.update(gff_update_iterator(db),merge_strategy="create_unique")
def plus_minus_strand(start,end):
    if start>end:
        return end,start
    else:
        return start,end
def segmentUnionLength(segment_list):
    points=[]
    for segment in segment_list:
        points.append((segment[0],False))
        points.append((segment[1],True))
    points_sorted = sorted(points, key=lambda x: int(x[0]))
    result=0
    counter=0
    for point_index in range(0,len(points_sorted)):
        if counter:
            result=result+(points_sorted[point_index][0]-points_sorted[point_index-1][0])
        if points_sorted[point_index][1]:
            counter=counter-1
        else:
            counter=counter+1
    return result
def alignment_segmentUnionLength(alignments):
    segment_list=[]
    for alignment in alignments:
        for hsp in alignment.hsps:
            if hsp.align_length < 100 and (hsp.identities/hsp.align_length) < 0.5: continue
            segment_list.append(plus_minus_strand(hsp.query_start,hsp.query_end))
    UnionLength=segmentUnionLength(segment_list)
    return UnionLength
def one_segment(query_length,hsp,species_count,gene_name):
    identity_discriminant_for_length=hsp.align_length/query_length
    identity_discriminant_for_identity_perscent=hsp.identities/hsp.align_length
    if hsp.align_length < 100 and identity_discriminant_for_length < 0.5:
        out_to_excel(species_count,gene_name,0)
    elif identity_discriminant_for_length==1 and identity_discriminant_for_identity_perscent==1:
        out_to_excel(species_count,gene_name,4)
    elif identity_discriminant_for_length==1 and hsp.gaps == 0:
        out_to_excel(species_count,gene_name,3)
    elif hsp.align_length>query_length or query_length-hsp.align_length <=50:
        out_to_excel(species_count,gene_name,2)
    else:
        out_to_excel(species_count,gene_name,1)
def multi_segment(query_length,UnionLength,species_count,gene_name):
    if UnionLength>query_length or query_length-UnionLength <=50:
        out_to_excel(species_count,gene_name,2)
    else:
        out_to_excel(species_count,gene_name,1)

def one_alignment(query_length,alignment,species_count,gene_name):
    if len(alignment.hsps)==1:
        one_segment(query_length,alignment.hsps[0],species_count,gene_name)
    else:
        UnionLength=alignment_segmentUnionLength([alignment])
        hsp_best=alignment.hsps[0]
        if UnionLength-hsp_best.align_length<50:
            one_segment(query_length,hsp_best,species_count,gene_name)
        else:
            multi_segment(query_length,UnionLength,species_count,gene_name)

def multi_alignment(query_length,alignments,species_count, gene_name):
    alignment_UnionLength_all=0
    alignment_list=[]
    for alignment in alignments:
        # 加这个if是因为，如果新添加的alignment没有使得原来的延长超过50bp，
        # 则添加新的没有意义，还不如不添加，还使得本可以是3或4的，变为1或2
        # 没有意义是指添加了overlap，没有增加有效长度
        if alignment_UnionLength_all > query_length-50:
            break
        alignment_list.append(alignment)
        alignment_UnionLength_all=alignment_segmentUnionLength(alignment_list)
    alignment_best=alignment_segmentUnionLength([alignments[0]])
    if alignment_UnionLength_all-alignment_best<50:
        one_alignment(query_length,alignments[0],species_count, gene_name)
    else:
        multi_segment(query_length,alignment_UnionLength_all,species_count,gene_name)
def excel(species_path,species_out_path,pav_excel_name):
    global pan_sh
    excel_book=Workbook()
    pan_sh=excel_book.active
    species_count=1
    for species_file in species_path.glob("*.xml"):
        # print (str(species_file)+"\n")
        species_count=species_count+1
        species_name=species_file.stem

        excel_species_name(species_count+1,species_name)
        with open(species_out_path/"xml"/(species_name+".xml")) as fl:
            for record in NCBIXML.parse(fl):
                gene_name=record.query.split()[0]
                if record.alignments:
                    #out_to_excel(species_count,record.query,1)
                    if len(record.alignments)==1:
                        one_alignment(record.query_length,record.alignments[0],species_count, gene_name)
                    else:
                        multi_alignment(record.query_length,record.alignments,species_count, gene_name)
                else:
                    out_to_excel(species_count,gene_name,0)
    excel_book.save(pav_excel_name)
def pav_contig_present_blast_gth_main (species_path,query_file_1,species_out_path_name,pav_excel_name):
    '''
    input:1 contig
    input 3 blastn query file,pan gene
    intermediate_out: species_out_path=Path('../Pan_genome_data/c_blast_present_contig/')
    output 1: pav_excel name
    '''
    global query_file
    query_file=query_file_1
    species_out_path=Path(species_out_path_name)
    
    global species_db_dir,species_out_xml_dir,species_out_asn_dir,species_out_txt_dir
    species_db_dir=directory_creater(species_out_path/"blastdb")
    species_out_xml_dir=directory_creater(species_out_path/"xml")
    species_out_asn_dir=directory_creater(species_out_path/"asn")
    species_out_txt_dir=directory_creater(species_out_path/"txt")
    species_out_gth_dir=directory_creater(species_out_path/"gth_out")
    species_out_gth_stdout_dir=directory_creater(species_out_path/"gth_stdout")
    gff_db_dir_path=directory_creater(species_out_path/"gff_db")
    # Parallel(n_jobs=12)(delayed(run_blast)(query_file,i,species_out_path) for i in species_path.glob("*.fasta"))
    # Parallel(n_jobs=12)(delayed(run_gth)(query_file,i,species_out_path) for i in species_path.glob("*.fasta"))
    # Parallel(n_jobs=12)(delayed(update_gff_db)(i,gff_db_dir_path) for i in species_out_gth_dir.glob("*.gff"))
    excel(species_out_xml_dir,species_out_path,pav_excel_name)

