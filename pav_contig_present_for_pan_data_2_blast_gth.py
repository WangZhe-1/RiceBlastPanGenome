#!/usr/bin/env python3
# -*- coding: utf-8 -*-
from ast import Num
from subprocess import STDOUT
from sys import stderr
from gffutils import attributes
from joblib import Parallel,delayed
from Directory_creater import directory_creater
from Bio.Blast.Applications import NcbiblastformatterCommandline, NcbimakeblastdbCommandline
from Bio.Blast.Applications import NcbiblastnCommandline
from Bio.Blast.Applications import NcbimakeblastdbCommandline
from Bio.Blast import NCBIXML
import os 
import re
from openpyxl import Workbook
from pathlib import Path
import subprocess
import gffutils

#output,remmben save at end

# excel_book=openpyxl.load_workbook('../test_why_00005_so_many_1/test.xlsx')

pan_sh=None
dic_gene={}
gene_count=1

def out_to_excel(count,gene,value):
    count=count+1
    global gene_count

    if dic_gene.__contains__(gene):
        pan_sh.cell(dic_gene[gene],count,value)
    else:
        gene_count=gene_count+1
        pan_sh.cell(gene_count,1,'Pan_{}'.format(gene_count-1))
        pan_sh.cell(gene_count,2,gene)

        dic_gene.setdefault(gene,gene_count)
        pan_sh.cell(dic_gene[gene],count,value)

def excel_species_name(count,name):
    pan_sh.cell(1,count,name)

def run_blast(query_file,species_id_path,species_out_path):
    make_db_cmd=NcbimakeblastdbCommandline(
        cmd='/mnt/d/zhes_learning_space/software_in_ubuntu/ncbi-blast-2.10.1+/bin/makeblastdb',
        dbtype='nucl',
        input_file=species_id_path,
        out=str(species_out_path/"blastdb"/species_id_path.stem)
    )
    blast_cmd=NcbiblastnCommandline(
            cmd='/mnt/d/zhes_learning_space/software_in_ubuntu/ncbi-blast-2.10.1+/bin/blastn',
            query=query_file,
            db=species_out_path/"blastdb"/species_id_path.stem,
            outfmt=11,
            out=species_out_path/"asn"/(species_id_path.stem+".asn")
            # perc_identity=95
        )
    blast_xml_cmd=NcbiblastformatterCommandline(
    archive=species_out_path/"asn"/(species_id_path.stem+".asn"),
    outfmt=5,
    out=species_out_path/"xml"/(species_id_path.stem+".xml"),
    cmd='/mnt/d/zhes_learning_space/software_in_ubuntu/ncbi-blast-2.10.1+/bin/blast_formatter'
    )
    blast_txt_cmd=NcbiblastformatterCommandline(
        archive=species_out_path/"asn"/(species_id_path.stem+".asn"),
        outfmt=7,
        out=species_out_path/"txt"/(species_id_path.stem+".txt"),
        cmd='/mnt/d/zhes_learning_space/software_in_ubuntu/ncbi-blast-2.10.1+/bin/blast_formatter'
    )
    db_file=species_out_path/"blastdb"/(species_id_path.stem+".ndb")
    if (species_out_path/"xml"/(species_id_path.stem+".xml")).exists() is False:
        if db_file.exists() is False:
            make_db_cmd()
        blast_cmd()
        blast_txt_cmd()
        blast_xml_cmd()
    # if os.path.getsize(str(species_out_path/"xml"/(species_id_path.stem+".xml"))) == 0:
    #     blast_cmd()
    #     blast_txt_cmd()
    #     blast_xml_cmd()

def run_gth(query_file,species_id_path,species_out_path):
    run_gth_popen=subprocess.Popen(
        [
            "/mnt/d/zhes_learning_space/software_in_ubuntu/gth-1.7.3-Linux_x86_64-64bit/bin/gth",
            "-genomic",
            species_id_path,
            "-cdna",
            query_file,
            "-gff3out",
            "-o",
            species_out_path/"gth_out"/(species_id_path.stem+".gff"),
            "-intermediate",
            "-force"
        ],
        stderr=subprocess.STDOUT,
        stdout=(species_out_path/"gth_stdout"/(species_id_path.stem+".txt")).open("w")
    )
    run_gth_popen.wait()

def update_gff_db(strain_gff,db_out_dir):
    def gff_update_iterator(db):
        for gene in db.features_of_type('gene'):

            # modify attributes

            # add a new attribute for exon id
            for target_id in gene.attributes["Target"][0].split(","):
                gene.attributes['ID'] = target_id.split(" ")[0]+"_"+gene.chrom
                gene.attributes['Target']=target_id
                yield gene

    try:
        gffutils.create_db(
        str(strain_gff),
        dbfn=str(db_out_dir/(strain_gff.stem+".db"))
        )
    except:
        pass
    db = gffutils.FeatureDB(str(db_out_dir/(strain_gff.stem+".db")))
    db.update(gff_update_iterator(db),merge_strategy="create_unique")

def one_alignment(record,alignments,species_count,gene_name,species_out_path,species_file):
    global max_flag
    for alignment in alignments:
        if len(alignment.hsps)==1:
            for hsp in alignment.hsps:
                identity_discriminant_for_length=hsp.align_length/record.query_length
                identity_discriminant_for_identity_perscent=hsp.identities/hsp.align_length
                max_flag=max_flag+2
                if hsp.align_length < 100 and identity_discriminant_for_length < 0.5:
                    out_to_excel(species_count,gene_name,0)
                elif identity_discriminant_for_length==1 and identity_discriminant_for_identity_perscent==1:
                    out_to_excel(species_count,gene_name,4)
                elif identity_discriminant_for_length==1 and hsp.gaps == 0:
                    out_to_excel(species_count,gene_name,3)
                elif hsp.align_length>record.query_length or record.query_length-hsp.align_length <=50:
                    out_to_excel(species_count,gene_name,2)
                else:
                    out_to_excel(species_count,gene_name,1)
        else:
            db = gffutils.FeatureDB(species_out_path/"gff_db"/(species_file.stem+".db"))
            try:
                gff_feature=db[gene_name+"_"+alignment.hit_def.split(" ")[0]]
                target_id=gff_feature.attributes["Target"][0]
                gth_length=abs(int(target_id.split(" ")[1])-int(target_id.split(" ")[2]))
                gth_identity=float(gff_feature.score)
            except gffutils.exceptions.FeatureNotFoundError:
                flag=0
                gth_length=0
                gth_identity=0
                while True:
                    try:
                        gff_feature=db[gene_name+"_"+alignment.hit_def.split(" ")[0]+str(flag)]
                        target_id=gff_feature.attributes["Target"][0]
                        gth_length_now=abs(int(target_id.split(" ")[1])-int(target_id.split(" ")[2]))
                        gth_length=max(gth_length,gth_length_now)
                        gth_identity=float(gff_feature.score)
                    except gffutils.exceptions.FeatureNotFoundError:
                        break
                #考虑使用gth就是不符合3，4的条件，所以不需要顾及是否identity，只考虑覆盖度的条件就好，这里就只考虑长度，选择长度长的。
            if gth_length < 100 and gth_identity< 0.5:
                out_to_excel(species_count,gene_name,0)
            elif gth_length>record.query_length or record.query_length-gth_length <=50:
                out_to_excel(species_count,gene_name,2)
            else:
                out_to_excel(species_count,gene_name,1)
def multi_alignment():
    pass
def excel(species_path,species_out_path,pav_excel_name):
    global pan_sh
    excel_book=Workbook()
    pan_sh=excel_book.active
    species_count=1
    for species_file in species_path.glob("ZJ2011-7-1.db"):
        # print (str(species_file)+"\n")
        species_count=species_count+1
        species_name=species_file.stem

        excel_species_name(species_count+1,species_name)
        with open(species_out_path/"xml"/(species_name+".xml")) as fl:
            for record in NCBIXML.parse(fl):
                gene_name=record.query.split()[0]
                if record.alignments:
                    global max_flag
                    max_flag=-1
                    #out_to_excel(species_count,record.query,1)
                    if max_flag>-1:continue
                    if len(record.alignments)==1:
                        one_alignment(record,record.alignments,species_count, gene_name, species_out_path, species_file)
                    else:
                        multi_alignment()
                else:
                    out_to_excel(species_count,gene_name,0)
    excel_book.save(pav_excel_name)
def pav_contig_present_blast_gth_main (species_path,query_file_1,species_out_path_name,pav_excel_name):
    '''
    input:1 contig
    input 3 blastn query file,pan gene
    intermediate_out: species_out_path=Path('../Pan_genome_data/c_blast_present_contig/')
    output 1: pav_excel name
    '''
    global query_file
    query_file=query_file_1
    species_out_path=Path(species_out_path_name)
    
    global species_db_dir,species_out_xml_dir,species_out_asn_dir,species_out_txt_dir
    species_db_dir=directory_creater(species_out_path/"blastdb")
    species_out_xml_dir=directory_creater(species_out_path/"xml")
    species_out_asn_dir=directory_creater(species_out_path/"asn")
    species_out_txt_dir=directory_creater(species_out_path/"txt")
    species_out_gth_dir=directory_creater(species_out_path/"gth_out")
    species_out_gth_stdout_dir=directory_creater(species_out_path/"gth_stdout")
    gff_db_dir_path=directory_creater(species_out_path/"gff_db")
    # Parallel(n_jobs=12)(delayed(run_blast)(query_file,i,species_out_path) for i in species_path.glob("*.fasta"))
    # Parallel(n_jobs=12)(delayed(run_gth)(query_file,i,species_out_path) for i in species_path.glob("*.fasta"))
    # Parallel(n_jobs=12)(delayed(update_gff_db)(i,gff_db_dir_path) for i in species_out_gth_dir.glob("*.gff"))
    excel(gff_db_dir_path,species_out_path,pav_excel_name)

