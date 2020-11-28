#!/usr/bin/env python3
# -*- coding: utf-8 -*-

from Bio.Blast.Applications import NcbimakeblastdbCommandline
from Bio.Blast.Applications import NcbiblastnCommandline
from Bio.Blast.Applications import NcbimakeblastdbCommandline
from Bio.Blast import NCBIXML
import os 
import re
import openpyxl
from openpyxl import Workbook
from pathlib import Path
from extract_strain_id import extract_strain_id

#output,remmben save at end

# excel_book=openpyxl.load_workbook('../test_why_00005_so_many_1/test.xlsx')

pan_sh=None
dic_gene={}
gene_count=1

def out_to_excel(count,gene,value):
    count=count+1
    global gene_count

    if dic_gene.__contains__(gene):
        pan_sh.cell(dic_gene[gene],count,value)
    else:
        gene_count=gene_count+1
        pan_sh.cell(gene_count,1,'Pan_{}'.format(gene_count-1))
        pan_sh.cell(gene_count,2,gene)

        dic_gene.setdefault(gene,gene_count)
        pan_sh.cell(dic_gene[gene],count,value)

def excel_species_name(count,name):
    pan_sh.cell(1,count,name)

def blastdb (db_file):
    make_db_cmd=NcbimakeblastdbCommandline(
        cmd='/mnt/d/zhes_learning_space/software_in_ubuntu/ncbi-blast-2.9.0+/bin/makeblastdb',
        dbtype='nucl',
        input_file=db_file
    )
    make_db_cmd()

def blast (species_path_name,busco_result,query_file,species_out_path_name,pav_excel_name):
    '''
    input:1 contig
    input 2 busco_result provide species list
    input 3 blastn query file,pan gene
    intermediate_out: species_out_path=Path('../Pan_genome_data/c_blast_present_contig/')
    output 1: pav_excel name
    '''
    global pan_sh
    species_path=Path(species_path_name)
    species_95_list=extract_strain_id(busco_result)
    species_95_list.append("70-15")
    species_95_list.append("ina168")
    species_95_list.remove("magnaporthe_oryzae_70-15_8_proteins_T0")
    species_out_path=Path(species_out_path_name)
    excel_book=Workbook()
    pan_sh=excel_book.active
    species_count=1
    
    for species_id in species_95_list:
        # print (species_id+"\t")
        for species_file in species_path.glob(species_id.strip('\n')+".fasta"):
            # print (str(species_file)+"\n")
            species_count=species_count+1
            species_name=species_file.stem

            excel_species_name(species_count+1,species_name)

            species_file_path=str(species_file)
            species_out=species_out_path/(species_name+'.xml')
            # species_out_1=os.path.join("../test_why_00005_so_many_1/",species_name+'.txt')
            # blastdb(species_file_path)
            blast_cmd=NcbiblastnCommandline(
                cmd='/mnt/d/zhes_learning_space/software_in_ubuntu/ncbi-blast-2.9.0+/bin/blastn',
                query=query_file,
                db=species_file_path,
                outfmt=5,
                out=species_out
                # perc_identity=95
            )
            if species_out.exists() is False:
                blast_cmd()
            if os.path.getsize(str(species_out)) == 0:
                blast_cmd()
            with open(species_out) as fl:
                for record in NCBIXML.parse(fl):
                    gene_name=record.query.split()[0]
                    if record.alignments:
                        max_flag=-1
                        #out_to_excel(species_count,record.query,1)
                        for alignment in record.alignments:
                            for hsp in alignment.hsps:
                                if max_flag == -1:
                                    identity_discriminant_for_length=hsp.align_length/record.query_length
                                    identity_discriminant_for_identity_perscent=hsp.identities/hsp.align_length
                                    max_flag=max_flag+2 
                                    if hsp.align_length < 100 and identity_discriminant_for_length < 0.5:
                                        out_to_excel(species_count,gene_name,0)
                                    elif identity_discriminant_for_length==1 and identity_discriminant_for_identity_perscent==1:
                                        out_to_excel(species_count,gene_name,4)
                                    elif identity_discriminant_for_length==1 and hsp.gaps == 0:
                                        out_to_excel(species_count,gene_name,3)
                                    elif hsp.align_length>record.query_length or record.query_length-hsp.align_length <=50:
                                        out_to_excel(species_count,gene_name,2)
                                    else:
                                        out_to_excel(species_count,gene_name,1)
                    else:
                        out_to_excel(species_count,gene_name,0)
    excel_book.save(pav_excel_name)

