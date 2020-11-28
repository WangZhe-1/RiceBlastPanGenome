#!/usr/bin/env python3
# -*- coding: utf-8 -*-

import openpyxl
from openpyxl import Workbook
from rpy2.robjects.packages import SignatureTranslatedAnonymousPackage
def set_minus(R_result):
    '''
    使用路径提供 1、2、3、4
    input 1: excel_book_add
    input 2: excel_book_minus    
    input 3: from_gene_file
    input 4: from_species_file
    output 1: set_minus_excel
    '''
    # 此文件的输入有4个

    #输入1和2
    excel_book_add=openpyxl.load_workbook(R_result+"set_minus_add_70-15.xlsx")
    excel_book_minus=openpyxl.load_workbook(R_result+"set_minus_minus_70-15.xlsx")
    # excel_book_add=openpyxl.load_workbook('../result/set_minus_add_Guy11.xlsx')
    # excel_book_minus=openpyxl.load_workbook('../result/set_minus_minus_Guy11.xlsx')

    pan_sh_add=excel_book_add.worksheets[0]
    pan_sh_minus=excel_book_minus.worksheets[0]
    out_excel=Workbook()
    pan_sh_out=out_excel.active

    # 输入3
    from_gene_file=R_result+"set_minus_gene_id_70-15.txt"
    # from_gene_file='../result/set_minus_gene_id_Guy11.txt'
    #输入4
    from_species_file=R_result+"set_minus_sort_protein_id_70-15.txt"
    # from_species_file='../result/set_minus_sort_protein_id_Guy11.txt'
    # from_species_file='../set_minus/test_0+1_input.txt'
    set_ori_plus=set()
    set_ori_minus=set()
    with open(from_gene_file,'r+') as fl:
        for line in fl:
            set_ori_plus.add(line.strip('\n'))
            set_ori_minus.add(line.strip('\n'))
    set_minus=set()
    set_plus=set()
    pan_sh_out.append(
        [
            "species",
            "minus",
            "add",
            "set_minus",
        ]
    )
    count=0
    count_min=0
    with open(from_species_file,'r+') as ffl:
        pan_sh_out.append(
        [
            str(ffl.readline().strip('\n')),
            len(set_ori_plus),
            len(set_ori_minus)
        ]
        )
        for line in ffl.readlines():
            set_minus.clear()
            set_plus.clear()
            ori_id=line.strip('\n').strip('\"')
            flag=-1
            # df[ori_id]
            # df1['category'] = np.where(df[ori_id] == 3 or df[ori_id] == 4, True, False)
            for i in range(pan_sh_add.min_column,pan_sh_add.max_column+1):
                if pan_sh_add.cell(1,i).value is None:continue
                if ori_id==str(pan_sh_add.cell(1,i).value):
                    # if str(ori_id).find(pan_sh_add.cell(1,i).value)!=-1:
                    specise_col=i
                    count=count+1
                    flag=flag+2
                    break
            if flag<0:
                print(ori_id)
                continue            
            for j in range(2,(pan_sh_minus.max_row+1)):
                cell_value=int(pan_sh_minus.cell(j,specise_col).value)
                if cell_value==0:
                    set_minus.add(pan_sh_minus.cell(j,1).value)
                    count_min=count_min+1
            for m in range(2,pan_sh_add.max_row+1):
                cell_value=int(pan_sh_add.cell(m,specise_col).value)
                if cell_value==1:
                    set_ori_plus.add(pan_sh_add.cell(m,1).value)
            set_ori_minus=set_ori_minus-set_minus
            pan_sh_out.append(
                [
                    str(ori_id),
                    len(set_ori_minus),
                    len(set_ori_plus),
                    len(set_minus)
                ]
            )
    print(count)
    print(count_min)
    # out_excel.save('../../wangzhe2/set_minus/set_minus/set_minus.xlsx')
    out_excel.save(R_result+"set_minus_set_minus.xlsx")
    # excel_book_add.save('../set_minus/test_0+1.xlsx')

def drawer_curve(R_result):
    '''
    input 1: set_minus_set_minus.xlsx
    output 1: set minus curve plot
    input and output are all in R_result
    '''    
    R_code_draw_curve='''
    drawer_curve=function(result_R){
    require(tidyverse)
    require(readxl)
    in_fl=read_xlsx(
        paste(result_R,"set_minus_set_minus.xlsx",sep = "")
        )
    in_fl$species=factor(in_fl$species,levels = in_fl$species)
    plot_line=ggplot(
        data = in_fl,
        mapping=aes(
        x=as.numeric(row.names(in_fl)),
        # y=minus
        group=1
        )
    )+
        geom_line(
        aes(
            y=minus
        ),
        size=1
        )+
        geom_line(
        aes(
            y=add
        ),
        size=1
        )+
        # scale_y_continuous(
        #   breaks = c(8511,9000,10000,11000,12000,12110),
        #   labels = c(8511,9000,10000,11000,12000,12110)
        # )+
        theme(
        panel.background = element_blank(),
        axis.line = element_line(size=1, colour = "black"),
        axis.title = element_text(size=20),
        axis.text = element_text(size = 20)
        )+
        xlab(label = "Number of strain")+
        ylab(label = "Number of gene")+
        geom_hline(
        yintercept = c(max(in_fl$add),min(in_fl$minus)),lty=3,lwd=1,alpha=0.8
        )+
        annotate(
        "text",
        x = 110,
        y = 9000,
        label="Core",
        size=10
        )+
        annotate(
        "text",
        x = 110,
        y = 14500,
        label="Pan",
        size=10
        )
    ggsave(
        paste(result_R,"curve.png",sep = ""),
        plot = plot_line,
        width = 8,
        height = 5
        )
    }
    '''
    R_darw_curve = SignatureTranslatedAnonymousPackage(R_code_draw_curve, "R_darw_curve")
    R_darw_curve.drawer_curve(R_result)